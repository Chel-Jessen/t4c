import time
import openai
import requests
import datetime
import csv
import openpyxl
from openpyxl.styles import PatternFill
from bs4 import BeautifulSoup
from urllib.parse import urlparse
from dotenv import load_dotenv
import prompts
from display_options import *

load_dotenv("openai.env")

openai.api_key = os.getenv("OPENAI_KEY")
openai.organization = os.getenv("OPENAI_ORG")


options = {
    "Titel": "",
    "Kurzbeschreibung": "",
    "Langbeschreibung": "",
    "Produktdetails": "",
    "Misc": []
}

csv_headers = []

data = []


def send_openai_request(key, value, input_text):
    prompt = ""
    match key:
        case "Titel":
            prompt = prompts.title_prompt.replace("%opt", value).replace("%text", input_text)
        case "Kurzbeschreibung":
            prompt = prompts.short_description_prompt.replace("%opt", value).replace("%text", input_text)
        case "Langbeschreibung":
            prompt = prompts.long_description_prompt.replace("%opt", value).replace("%text", input_text)
        case "Produktdetails":
            prompt = prompts.product_details_prompt.replace("%opt", value).replace("%text", input_text)
        case "Misc":
            match value:
                case "Artikelnummer":
                    prompt = prompts.article_number_prompt.replace("%opt", value).replace("%text", input_text)
                case "Meta-Beschreibung in 2 Sätzen":
                    prompt = prompts.meta_description_prompt.replace("%opt", value).replace("%text", input_text)
                case "4 Meta-Keywords in Kombination mit den Wörtern B2B und Dropshipping":
                    prompt = prompts.meta_keywords_prompt.replace("%opt", value).replace("%text", input_text)
                case "20 Keywords":
                    prompt = prompts.keywords_prompt.replace("%opt", value).replace("%text", input_text)
                case "Kategorie":
                    prompt = prompts.category_prompt.replace("%opt", value).replace("%text", input_text)
    return reqopenai(prompt, key)


def reqopenai(prompt, key):
    max_tokens = 100
    try:
        match key:
            case "Titel":
                max_tokens = os.getenv("MAX_TOKENS_TITLE")
            case "Kurzbeschreibung":
                max_tokens = os.getenv("MAX_TOKENS_SHORTDESCRIPTION")
            case "Langbeschreibung":
                max_tokens = os.getenv("MAX_TOKENS_LONGDESCRIPTION")
            case "Produktdetails":
                max_tokens = os.getenv("MAX_TOKENS_PRODUCTDETAILS")
            case "Misc":
                max_tokens = os.getenv("MAX_TOKENS_MISC")
        completion = openai.Completion.create(model="text-davinci-003", prompt=prompt, max_tokens=max_tokens)
        return completion.choices[0].text.strip()
    except openai.error.APIError:
        return "An `APIError` indicates that something went wrong on our side when processing your request. This could be due to a temporary error, a bug, or a system outage."
    except openai.error.APIConnectionError:
        return "An `APIConnectionError` indicates that your request could not reach our servers or establish a secure connection. This could be due to a network issue, a proxy configuration, an SSL certificate, or a firewall rule."
    except openai.error.RateLimitError:
        return "A `RateLimitError` indicates that you have hit your assigned rate limit. This means that you have sent too many tokens or requests in a given period of time, and our services have temporarily blocked you from sending more."
    except openai.error.ServiceUnavailableError:
        return "A `ServiceUnavailableError` indicates that our servers are temporarily unable to handle your request. This could be due to a planned or unplanned maintenance, a system upgrade, or a server failure. These errors can also be returned during periods of high traffic."
    except openai.error.InvalidRequestError:
        return "An InvalidRequestError indicates that your request was malformed or exceeded the maximum tokens supported by the model."


def display():
    global options
    clear()
    choice = display_options("main")
    match choice:
        case "1":
            dc = display_options("Titel")
            if dc != "0":
                options["Titel"] = dc
            else:
                options["Titel"] = ""
            display()
        case "2":
            dc = display_options("Kurzbeschreibung")
            if dc != "0":
                options["Kurzbeschreibung"] = dc
            else:
                options["Kurzbeschreibung"] = ""
            display()
        case "3":
            dc = display_options("Langbeschreibung")
            if dc != "0":
                options["Langbeschreibung"] = dc
            else:
                options["Langbeschreibung"] = ""
            display()
        case "4":
            dc = display_options("Produktdetails")
            if dc != "0":
                options["Produktdetails"] = dc
            else:
                options["Produktdetails"] = ""
            display()
        case "5":
            dc = display_options("Misc")
            if dc != "0":
                options["Misc"].append(dc)
            else:
                options["Misc"] = []
            display()
        case "6":
            preset_choice = display_presets("Voreingestellt")
            try:
                options = opts["Voreingestellt"][preset_choice]
                display()
            except KeyError:
                clear()
                print("Ungültiger Wert")
                display()
        case "0":
            clear()
            if display_validate(options):
                format_options()
                get_csv_headers()
            else:
                display()
        case _:
            print("Ungültige Eingabe!")
            display()


def format_options():
    global options
    copy_options = options.copy()
    for key, value in copy_options.items():
        if isinstance(value, str) and value.strip() == "":
            del options[key]
        elif isinstance(value, list) and not value:
            del options[key]
        elif key in opts:
            if key == "Misc":
                for i, sub_key in enumerate(copy_options[key]):
                    options[key][i] = opts[key][sub_key]
                continue
            options[key] = opts[key][value]
        else:
            print("Format Fehler!")
            exit()


def get_csv_headers():
    for value in options.values():
        if isinstance(value, list):
            csv_headers.extend(value)
        else:
            csv_headers.append(value)


def is_url(string):
    try:
        result = urlparse(string)
        return all([result.scheme, result.netloc])
    except:
        return False


def get_parent_and_subkey(csv_header: str):
    for key, sub_dict in opts.items():
        if key not in ["main", "Voreingestellt"]:
            for subkey, val in sub_dict.items():
                if val == csv_header:
                    return key, subkey
    print("Subkey konnte nicht gefunden werden")
    exit()


def read_file(path: str):
    if not os.path.exists(path) or not os.path.isfile(path):
        print("Dieser Pfad existiert nicht, oder ist keine Datei!")
        exit()
    if path.endswith('.csv'):
        col_name = input("In welcher Spalte befindet sich der Link? ")
        with open(path, 'r', encoding="UTF-8") as file:
            reader = csv.reader(file, delimiter=";")
            csv_data = list(reader)
            header = csv_data[0]
            header_indices = {header[i]: i for i in range(len(header))}
            try:
                index = header_indices[col_name]
                return [row[index] for row in csv_data[1:] if row[index]]
            except KeyError:
                print("Diese Spalte existiert nicht!")
                exit()
            except IndexError:
                print("Diese Spalte existiert nicht!")
                exit()
    elif path.endswith('.xlsx'):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        col_name = input("In welcher Spalte befindet sich der Link? ")
        if not col_name.isnumeric():
            print("Bitte Spalten nur als Nummern eingeben")
            exit()
        column = []
        for row in sheet.iter_rows(min_col=int(col_name), max_col=int(col_name), values_only=True):
            column.append(row[0])
        if all(item is None for item in column):
            print("Keine Werte gefunden! Haben Sie vielleicht die falsche Spalte?")
            exit()
        return column
    else:
        print('Unbekannter Dateityp. Bitte nur .csv oder .xlsx Dateien')
        exit()


def write_file(urls, outfile=""):
    if outfile == "":
        outfile = f"output_{datetime.datetime.now().strftime('%Y-%m-%d %H-%M-%S').replace(':', '-')}.xlsx"
    if not outfile.endswith(".xlsx"):
        outfile += ".xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.cell(row=1, column=1, value="URL")
    for i, url in enumerate(urls, start=2):
        sheet.cell(row=i, column=1, value=url)
    # write csv_headers to xlsx
    for col, val in enumerate(csv_headers, start=2):
        sheet.cell(row=1, column=col, value=val)
    # write the openai_data in xlsx
    for row, item in enumerate(data, start=2):
        for col, item_val in enumerate(item, start=2):
            check_value(row, col, item_val, sheet)
            sheet.cell(row=row, column=col, value=item_val)
    workbook.save(outfile)


def check_value(row: int, col: int, item, worksheet):
    # declare all sensitive values
    max_char_sensitive = [("Titel", "1", 50), ("Titel", "2", 70), ("Titel", "3", 80), ("Titel", "4", 120),
                          ("Titel", "5", 132), ("Titel", "6", 140), ("Titel", "7", 150), ("Titel", "8", 200),
                          ("Titel", "10", 100),
                          ("Kurzbeschreibung", "1", 140), ("Kurzbeschreibung", "2", 150),
                          ("Kurzbeschreibung", "3", 250), ("Kurzbeschreibung", "4", 1000),
                          ("Langbeschreibung", "1", 2000), ("Langbeschreibung", "2", 2000),
                          ("Langbeschreibung", "3", 4000), ("Langbeschreibung", "4", 4000),
                          ("Langbeschreibung", "5", 5000), ("Langbeschreibung", "6", 5000),
                          ("Langbeschreibung", "7", 5000),
                          ("Produktdetails", "1", 250), ("Produktdetails", "3", 900), ("Produktdetails", "4", 2000)]
    html_sensitive = [("Kurzbeschreibung", "4"), ("Kurzbeschreibung", "7"), ("Langbeschreibung", "2"),
                      ("Langbeschreibung", "4"), ("Langbeschreibung", "7")]
    frz_sensitive = [("Titel", "5"), ("Kurzbeschreibung", "3"), ("Langbeschreibung", "6")]
    max_words_sensitive = [("Produktdetails", "2", 25), ("Misc", "3", 4), ("Misc", "4", 20)]
    others_sensitive = [("Titel", "2"), ("Titel", "10"), ("Titel", "11"), ("Kurzbeschreibung", "6"),
                        ("Kurzbeschreibung", "7"), ("Misc", "1"), ("Misc", "2"), ("Misc", "5"), ("Misc", "6")]

    # set the PatterFill colors
    red = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Max_Chars
    yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # HTML
    blue = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # FRZ
    pink = PatternFill(start_color='FFC0CB', end_color='FFC0CB', fill_type='solid')  # Max_Words
    brown = PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')  # Others
    black = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    # get the corresponding csv_header index
    csv_header = csv_headers[col - 2]

    # get parent and sub key from opts and the csv_header
    parent_key, local_key = get_parent_and_subkey(csv_header)
    # check for max_chars
    for check in max_char_sensitive:
        if check[0] == parent_key and check[1] == local_key:
            max_chars = check[2]
            if len(item) > max_chars:
                worksheet.cell(row=row, column=col).fill = red
            break

    # check for HTML
    for check in html_sensitive:
        if check[0] == parent_key and check[1] == local_key:
            worksheet.cell(row=row, column=col).fill = yellow
            break

    # check for FRZ
    for check in frz_sensitive:
        if check[0] == parent_key and check[1] == local_key:
            worksheet.cell(row=row, column=col).fill = blue
            break

    # check for max_words
    for check in max_words_sensitive:
        if check[0] == parent_key and check[1] == local_key:
            if len(item.split(" ")) + 1 > check[2]:
                worksheet.cell(row=row, column=col).fill = pink
            break

    # check for others
    for check in others_sensitive:
        if check[0] == parent_key and check[1] == local_key:
            worksheet.cell(row=row, column=col).fill = brown
            break

    if "indicates that" in item:
        worksheet.cell(row=row, column=col).fill = black


def fetch(url: str):
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, "html.parser")
    cleantext = ' '.join(soup.text.strip().split())
    return cleantext


def main():
    file_path = input("Bitte geben Sie den Dateipfad ein: ")
    urls = read_file(file_path)
    urls = [url for url in urls if is_url(url)]
    display()
    t_start = time.time()
    print("URLs werden eingelesen...")
    url_len = len(urls)
    for i, url in enumerate(urls, start=1):
        print(f"[{i}/{url_len}]  URL:{url} wird gescraped...")
        text = fetch(url)
        results = []
        for key, val in options.items():
            if isinstance(val, list):
                for item in val:
                    print(f"[{i}/{url_len}]  {item} wird generiert...")
                    results.append(send_openai_request(key, item, text))
            else:
                print(f"[{i}/{url_len}]  {key} wird generiert...")
                results.append(send_openai_request(key, val, text))
        data.append(results)
    print("")
    t_end = time.time()
    outfile_name = input("Output-Dateiname: ")
    write_file(urls, outfile_name)
    print(f"Dauer: {(t_end - t_start):0.f} Sekunden")


if __name__ == '__main__':
    main()
    print("Fertig :)")
